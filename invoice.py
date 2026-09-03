import streamlit as st
import io
import re
import html
import unicodedata

from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import streamlit.components.v1 as components
from openpyxl import load_workbook, Workbook


# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="Bill / Invoice Search",
    page_icon="🔎",
    layout="wide"
)


# ============================================================
# TITLE
# ============================================================

st.title("🔎 Bill / Invoice Search")

st.caption(
    "Search Bill / Invoice numbers across all uploaded Excel files."
)


# ============================================================
# BASIC FUNCTIONS
# ============================================================

def is_blank(value):

    if value is None:
        return True

    try:
        if value != value:
            return True
    except Exception:
        pass

    return str(value).strip() == ""


def normalize_value(value):

    if is_blank(value):
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):

        if value.is_integer():
            return str(int(value))

        return str(value).strip()

    return str(value).strip()


def normalize_search_value(value):

    if is_blank(value):
        return ""

    text = normalize_value(value)

    text = unicodedata.normalize(
        "NFKC",
        text
    )

    text = text.replace(
        "\xa0",
        " "
    )

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip().lower()


# ============================================================
# NORMALIZE HEADER
# ============================================================

def normalize_header(value):

    if is_blank(value):
        return ""

    text = unicodedata.normalize(
        "NFKC",
        str(value)
    )

    text = text.strip().lower()

    text = re.sub(
        r"[_\-/\\.:]+",
        " ",
        text
    )

    text = re.sub(
        r"[^\w\s]",
        " ",
        text
    )

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip()


# ============================================================
# BILL / INVOICE HEADER DETECTION
# ============================================================

def is_bill_invoice_header(value):

    header = normalize_header(value)

    if not header:
        return False

    exact_names = {

        "bill",
        "bill no",
        "bill number",
        "bill num",

        "bil",
        "bil no",
        "bil number",
        "bil num",

        "invoice",
        "invoice no",
        "invoice number",
        "invoice num",

        "inovice",
        "inovice no",
        "inovice number",
        "inovice num",

        "inv",
        "inv no",
        "inv number",
        "inv num",

        "bill ref",
        "bill ref no",
        "bill ref number",

        "invoice ref",
        "invoice ref no",
        "invoice ref number",

        "inovice ref",
        "inovice ref no",
        "inovice ref number"
    }

    if header in exact_names:
        return True

    words = set(
        header.split()
    )

    bill_words = {
        "bill",
        "bil",
        "invoice",
        "inovice",
        "inv"
    }

    number_words = {
        "no",
        "number",
        "num"
    }

    if (
        words.intersection(bill_words)
        and
        words.intersection(number_words)
    ):
        return True

    return False


# ============================================================
# SERIAL NUMBER HEADER
# ============================================================

def is_serial_header(value):

    text = normalize_header(value)

    return text in {

        "s no",
        "sno",

        "s number",
        "serial",
        "serial no",
        "serial number",

        "sl no",
        "slno",
        "sl number"
    }


# ============================================================
# FIND ALL BILL / INVOICE HEADERS
# ============================================================

def find_bill_invoice_headers(rows):

    headers = []

    for row_index, row in enumerate(rows):

        for column_index, value in enumerate(row):

            if is_bill_invoice_header(value):

                headers.append({

                    "row":
                        row_index,

                    "column":
                        column_index,

                    "value":
                        value
                })

    return headers


# ============================================================
# EXCEL COLUMN LETTER
# ============================================================

def excel_column_letter(column_number):

    result = ""

    while column_number > 0:

        column_number, remainder = divmod(
            column_number - 1,
            26
        )

        result = (
            chr(65 + remainder)
            + result
        )

    return result


# ============================================================
# FIND HEADER BLOCK START
#
# This is specifically designed for side-by-side tables.
#
# Example:
#
# A       B           C       D       E...
# S NO | INVOICE NO | SHOP ID | ...
#
# invoice_column = B
#
# Start must become A.
#
#
# Example with previous side table:
#
# A ... H | blank | J       K           L...
#                  S NO   | INVOICE NO | SHOP ID
#
# invoice_column = K
#
# Start must become J, NOT A.
# ============================================================

def find_horizontal_start(
    rows,
    header_row,
    invoice_column,
    all_headers
):

    row = list(
        rows[header_row]
    )

    # --------------------------------------------------------
    # FIRST:
    # Find previous invoice header ON THE SAME ROW.
    # --------------------------------------------------------

    same_row_previous_headers = [

        item["column"]

        for item in all_headers

        if (
            item["row"] == header_row
            and
            item["column"] < invoice_column
        )
    ]

    if same_row_previous_headers:

        previous_invoice_column = max(
            same_row_previous_headers
        )

        # ----------------------------------------------------
        # Current table must start AFTER previous table.
        #
        # Search from current invoice column leftward.
        # Prefer S NO / SNO.
        # ----------------------------------------------------

        for col in range(
            invoice_column - 1,
            previous_invoice_column,
            -1
        ):

            if col < len(row):

                if is_serial_header(
                    row[col]
                ):

                    return col

        # ----------------------------------------------------
        # If no SNO found, find the first nonblank column
        # after a blank separator.
        # ----------------------------------------------------

        start = invoice_column

        for col in range(
            invoice_column - 1,
            previous_invoice_column,
            -1
        ):

            if col >= len(row):
                continue

            if is_blank(
                row[col]
            ):

                break

            start = col

        return start

    # --------------------------------------------------------
    # NO previous invoice header on same row.
    #
    # Prefer S NO before invoice column.
    # --------------------------------------------------------

    for col in range(
        invoice_column - 1,
        -1,
        -1
    ):

        if col < len(row):

            if is_serial_header(
                row[col]
            ):

                return col

    # --------------------------------------------------------
    # Otherwise search left until separator.
    # --------------------------------------------------------

    start = invoice_column

    blank_count = 0

    for col in range(
        invoice_column - 1,
        -1,
        -1
    ):

        if col >= len(row):
            continue

        value = row[col]

        if is_blank(value):

            blank_count += 1

            if blank_count >= 1:
                break

        else:

            start = col
            blank_count = 0

    return start


# ============================================================
# FIND NEXT TABLE START ON SAME HEADER ROW
#
# This is the MAIN FIX.
#
# If:
#
# A B C D E F G H | I | J K L M...
#
# first invoice = B
# second invoice = K
#
# We locate the S NO immediately before K.
# That gives J.
#
# Therefore first table ends at J.
# Python slice A:J excludes J.
# ============================================================

def find_next_table_start(
    rows,
    header_row,
    invoice_column,
    all_headers
):

    row = list(
        rows[header_row]
    )

    next_headers = sorted([

        item

        for item in all_headers

        if (
            item["row"] == header_row
            and
            item["column"] > invoice_column
        )

    ], key=lambda item: item["column"])

    if not next_headers:

        return None

    next_invoice = next_headers[0][
        "column"
    ]

    # --------------------------------------------------------
    # Search LEFT from next INVOICE header.
    #
    # Usually:
    #
    # S NO | INVOICE NO
    #
    # so this finds S NO.
    # --------------------------------------------------------

    for col in range(
        next_invoice - 1,
        invoice_column,
        -1
    ):

        if col >= len(row):
            continue

        if is_serial_header(
            row[col]
        ):

            return col

    # --------------------------------------------------------
    # If S NO isn't present, locate first column of the
    # next contiguous header block.
    # --------------------------------------------------------

    next_start = next_invoice

    for col in range(
        next_invoice - 1,
        invoice_column,
        -1
    ):

        if col >= len(row):
            continue

        if is_blank(
            row[col]
        ):

            break

        next_start = col

    return next_start


# ============================================================
# FIND NATURAL RIGHT EDGE
#
# Used when there is NO second table on the same header row.
#
# IMPORTANT:
# Blank HEADER cells alone do NOT stop the table because
# screenshot 1 contains valid data below blank headers.
# ============================================================

def find_natural_table_end(
    rows,
    header_row,
    start_column,
    invoice_column,
    end_row
):

    if not rows:
        return invoice_column + 1

    sheet_width = max(
        len(row)
        for row in rows
    )

    last_used = invoice_column

    # --------------------------------------------------------
    # Limit the scan to this vertical table area.
    # --------------------------------------------------------

    for row_number in range(
        header_row,
        end_row
    ):

        row = rows[row_number]

        for col in range(
            start_column,
            len(row)
        ):

            if not is_blank(
                row[col]
            ):

                last_used = max(
                    last_used,
                    col
                )

    return min(
        last_used + 1,
        sheet_width
    )


# ============================================================
# FIND VERTICAL TABLE END
#
# IMPORTANT:
# Ignore side-by-side invoice headers on SAME row.
#
# Only later ROWS can terminate the vertical table.
# ============================================================

def find_vertical_table_end(
    rows,
    header_row,
    invoice_column,
    all_headers
):

    total_rows = len(
        rows
    )

    candidates = []

    for item in all_headers:

        if item["row"] <= header_row:
            continue

        # ----------------------------------------------------
        # A later invoice header is considered a new vertical
        # table only if it is reasonably near the current
        # invoice column.
        # ----------------------------------------------------

        if abs(
            item["column"] -
            invoice_column
        ) <= 3:

            candidates.append(
                item["row"]
            )

    if candidates:

        return min(
            candidates
        )

    return total_rows


# ============================================================
# CHECK WHETHER ROW HAS ANY DATA INSIDE TABLE
# ============================================================

def row_has_table_data(
    row,
    start_column,
    end_column
):

    for col in range(
        start_column,
        end_column
    ):

        if col < len(row):

            if not is_blank(
                row[col]
            ):

                return True

    return False


# ============================================================
# IMPROVE / RECONSTRUCT HEADERS
#
# Fixes screenshot 1.
#
# IMPORTANT:
# Existing column names are kept exactly.
#
# For blank headers:
# 1. Look at nearby rows ABOVE.
# 2. Do NOT use numeric data rows as headers.
# 3. If nothing is available, show Excel column name.
# ============================================================

def improve_headers(
    rows,
    header_row,
    start_column,
    end_column
):

    source_header = list(
        rows[header_row]
    )

    while len(
        source_header
    ) < end_column:

        source_header.append(
            None
        )

    headers = source_header[
        start_column:end_column
    ]

    # --------------------------------------------------------
    # Only look ABOVE first.
    #
    # Looking below can accidentally use data values as
    # column names.
    # --------------------------------------------------------

    candidate_rows = []

    for distance in range(
        1,
        8
    ):

        candidate = (
            header_row -
            distance
        )

        if candidate >= 0:

            candidate_rows.append(
                candidate
            )

    for index in range(
        len(headers)
    ):

        if not is_blank(
            headers[index]
        ):
            continue

        absolute_column = (
            start_column +
            index
        )

        replacement = None

        for candidate_row in candidate_rows:

            row = rows[
                candidate_row
            ]

            if absolute_column >= len(
                row
            ):
                continue

            value = row[
                absolute_column
            ]

            if is_blank(
                value
            ):
                continue

            text = str(
                value
            ).strip()

            # ------------------------------------------------
            # Reject plain numbers.
            # ------------------------------------------------

            if re.fullmatch(
                r"[-+]?\d+(?:\.\d+)?",
                text
            ):

                continue

            # ------------------------------------------------
            # Reject date-like objects.
            # ------------------------------------------------

            if hasattr(
                value,
                "year"
            ) and hasattr(
                value,
                "month"
            ):

                continue

            # ------------------------------------------------
            # Header should contain some alphabetic text.
            # ------------------------------------------------

            if not re.search(
                r"[A-Za-z]",
                text
            ):

                continue

            replacement = value

            break

        if replacement is not None:

            headers[
                index
            ] = replacement

    # --------------------------------------------------------
    # Still blank:
    # show Excel column letter instead of empty heading.
    # --------------------------------------------------------

    for index in range(
        len(headers)
    ):

        if is_blank(
            headers[index]
        ):

            absolute_excel_column = (
                start_column +
                index +
                1
            )

            headers[
                index
            ] = (
                "Column "
                + excel_column_letter(
                    absolute_excel_column
                )
            )

    return headers


# ============================================================
# BUILD ONE TABLE
# ============================================================

def build_table(
    rows,
    filename,
    sheet_name,
    header_info,
    all_headers
):

    header_row = header_info[
        "row"
    ]

    invoice_column = header_info[
        "column"
    ]

    invoice_header = header_info[
        "value"
    ]

    # ========================================================
    # LEFT BOUNDARY
    # ========================================================

    start_column = find_horizontal_start(
        rows,
        header_row,
        invoice_column,
        all_headers
    )

    # ========================================================
    # VERTICAL END
    # ========================================================

    end_row = find_vertical_table_end(
        rows,
        header_row,
        invoice_column,
        all_headers
    )

    # ========================================================
    # RIGHT BOUNDARY
    #
    # FIRST priority:
    # next table on SAME header row.
    # ========================================================

    next_table_start = find_next_table_start(
        rows,
        header_row,
        invoice_column,
        all_headers
    )

    if next_table_start is not None:

        end_column = next_table_start

    else:

        end_column = find_natural_table_end(
            rows,
            header_row,
            start_column,
            invoice_column,
            end_row
        )

    # ========================================================
    # SAFETY
    # ========================================================

    if start_column < 0:

        start_column = 0

    if start_column > invoice_column:

        start_column = invoice_column

    if end_column <= invoice_column:

        end_column = (
            invoice_column +
            1
        )

    # ========================================================
    # HEADERS
    # ========================================================

    headers = improve_headers(
        rows,
        header_row,
        start_column,
        end_column
    )

    if not headers:

        return None

    # ========================================================
    # RELATIVE INVOICE COLUMN
    # ========================================================

    relative_invoice_column = (
        invoice_column -
        start_column
    )

    if (
        relative_invoice_column < 0
        or
        relative_invoice_column >= len(headers)
    ):

        return None

    # ========================================================
    # SEARCH INDEX
    # ========================================================

    search_index = defaultdict(
        list
    )

    # ========================================================
    # DATA ROWS
    # ========================================================

    for row_number in range(
        header_row + 1,
        end_row
    ):

        source_row = list(
            rows[row_number]
        )

        # ----------------------------------------------------
        # Skip completely blank rows inside this table.
        # ----------------------------------------------------

        if not row_has_table_data(
            source_row,
            start_column,
            end_column
        ):

            continue

        # ----------------------------------------------------
        # Extend only to THIS table's end.
        # ----------------------------------------------------

        while len(
            source_row
        ) < end_column:

            source_row.append(
                None
            )

        # ----------------------------------------------------
        # CRITICAL:
        #
        # Slice ONLY current table.
        #
        # Side-by-side table is NOT included.
        # ----------------------------------------------------

        table_row = source_row[
            start_column:end_column
        ]

        # ----------------------------------------------------
        # Exact width
        # ----------------------------------------------------

        if len(
            table_row
        ) < len(headers):

            table_row.extend(

                [None] * (
                    len(headers) -
                    len(table_row)
                )
            )

        table_row = table_row[
            :len(headers)
        ]

        # ----------------------------------------------------
        # Invoice value
        # ----------------------------------------------------

        invoice_value = table_row[
            relative_invoice_column
        ]

        normalized = normalize_search_value(
            invoice_value
        )

        if not normalized:
            continue

        search_index[
            normalized
        ].append({

            "excel_row":
                row_number + 1,

            "row":
                table_row
        })

    return {

        "filename":
            filename,

        "sheet":
            sheet_name,

        "header":
            invoice_header,

        "header_row":
            header_row + 1,

        "invoice_column":
            invoice_column + 1,

        "start_column":
            start_column + 1,

        "end_column":
            end_column,

        "headers":
            headers,

        "search_index":
            dict(
                search_index
            )
    }


# ============================================================
# READ ONE EXCEL FILE
# ============================================================

def read_excel_file(
    filename,
    file_bytes
):

    result = {

        "filename":
            filename,

        "tables":
            [],

        "error":
            None
    }

    try:

        workbook = load_workbook(
            filename=io.BytesIO(
                file_bytes
            ),
            read_only=True,
            data_only=True
        )

        for worksheet in workbook.worksheets:

            rows = []

            for row in worksheet.iter_rows(
                values_only=True
            ):

                rows.append(
                    list(row)
                )

            if not rows:
                continue

            headers = find_bill_invoice_headers(
                rows
            )

            if not headers:
                continue

            for header_info in headers:

                table = build_table(
                    rows,
                    filename,
                    worksheet.title,
                    header_info,
                    headers
                )

                if table is not None:

                    result[
                        "tables"
                    ].append(
                        table
                    )

        workbook.close()

    except Exception as error:

        result[
            "error"
        ] = str(
            error
        )

    return result


# ============================================================
# BUILD SEARCH INDEX
# ============================================================

@st.cache_data(
    show_spinner=False,
    max_entries=5
)
def build_search_index(
    file_data
):

    results = []

    worker_count = min(
        4,
        max(
            1,
            len(file_data)
        )
    )

    with ThreadPoolExecutor(
        max_workers=worker_count
    ) as executor:

        futures = {}

        for filename, file_bytes in file_data:

            future = executor.submit(
                read_excel_file,
                filename,
                file_bytes
            )

            futures[
                future
            ] = filename

        for future in as_completed(
            futures
        ):

            results.append(
                future.result()
            )

    results.sort(
        key=lambda x:
            x["filename"].lower()
    )

    return results


# ============================================================
# SEARCH
# ============================================================

def search_excel(
    indexed_files,
    search_value
):

    search_value = normalize_search_value(
        search_value
    )

    results = []

    seen = set()

    for file_data in indexed_files:

        for table in file_data[
            "tables"
        ]:

            matches = table[
                "search_index"
            ].get(
                search_value,
                []
            )

            for match in matches:

                unique_key = (

                    table["filename"],

                    table["sheet"],

                    table["header_row"],

                    table["invoice_column"],

                    match["excel_row"]
                )

                if unique_key in seen:

                    continue

                seen.add(
                    unique_key
                )

                results.append({

                    "filename":
                        table["filename"],

                    "sheet":
                        table["sheet"],

                    "header":
                        table["header"],

                    "header_row":
                        table["header_row"],

                    "invoice_column":
                        table["invoice_column"],

                    "start_column":
                        table["start_column"],

                    "end_column":
                        table["end_column"],

                    "excel_row":
                        match["excel_row"],

                    "headers":
                        table["headers"],

                    "row":
                        match["row"]
                })

    return results


# ============================================================
# SAFE HTML
# ============================================================

def safe_html(value):

    if is_blank(value):
        return ""

    return html.escape(
        str(value)
    )


# ============================================================
# DISPLAY TABLE
# ============================================================

def display_excel_table(
    headers,
    rows,
    height=650
):

    headers = list(
        headers
    )

    rows = [

        list(row)

        for row in rows
    ]

    maximum_columns = max(

        [len(headers)] +

        [len(row) for row in rows] +

        [1]
    )

    # --------------------------------------------------------
    # Headers
    # --------------------------------------------------------

    while len(
        headers
    ) < maximum_columns:

        headers.append(
            f"Column {len(headers) + 1}"
        )

    # --------------------------------------------------------
    # Rows
    # --------------------------------------------------------

    fixed_rows = []

    for row in rows:

        if len(
            row
        ) < maximum_columns:

            row.extend(

                [None] * (
                    maximum_columns -
                    len(row)
                )
            )

        fixed_rows.append(
            row[:maximum_columns]
        )

    rows = fixed_rows

    # ========================================================
    # HTML
    # ========================================================

    table_html = """
<!DOCTYPE html>

<html>

<head>

<meta charset="UTF-8">

<style>

html,
body {

    margin: 0;
    padding: 0;

    background: transparent;

    font-family:
        -apple-system,
        BlinkMacSystemFont,
        "Segoe UI",
        sans-serif;
}

.wrapper {

    width: 100%;

    max-height: 650px;

    overflow-x: auto;
    overflow-y: auto;

    border: 1px solid #30333d;

    border-radius: 8px;
}

table {

    border-collapse: collapse;

    width: max-content;

    min-width: 100%;

    table-layout: auto;

    font-size: 14px;
}

th {

    position: sticky;

    top: 0;

    z-index: 50;

    background: #1f2129;

    color: #e5e7eb;

    border: 1px solid #3a3d47;

    padding: 11px 14px;

    text-align: left;

    white-space: nowrap;

    font-weight: 600;

    min-width: 65px;
}

td {

    background: #0e1117;

    color: #f5f5f5;

    border: 1px solid #30333d;

    padding: 10px 14px;

    text-align: left;

    white-space: nowrap;

    min-width: 65px;
}

tr:hover td {

    background: #181b23;
}

::-webkit-scrollbar {

    width: 12px;
    height: 12px;
}

::-webkit-scrollbar-track {

    background: #16181f;
}

::-webkit-scrollbar-thumb {

    background: #777;

    border-radius: 8px;
}

::-webkit-scrollbar-thumb:hover {

    background: #999;
}

</style>

</head>

<body>

<div class="wrapper">

<table>

<thead>

<tr>
"""

    for header in headers:

        table_html += (
            "<th>"
            + safe_html(header)
            + "</th>"
        )

    table_html += """

</tr>

</thead>

<tbody>
"""

    for row in rows:

        table_html += "<tr>"

        for value in row[
            :maximum_columns
        ]:

            table_html += (
                "<td>"
                + safe_html(value)
                + "</td>"
            )

        table_html += "</tr>"

    table_html += """

</tbody>

</table>

</div>

</body>

</html>
"""

    calculated_height = min(
        height,
        max(
            220,
            100 + len(rows) * 43
        )
    )

    components.html(
        table_html,
        height=calculated_height,
        scrolling=False
    )


# ============================================================
# DETECTED HEADERS DISPLAY
# ============================================================

def display_detected_headers(
    indexed_files
):

    headers = [

        "File",

        "Sheet",

        "Bill / Invoice Header",

        "Excel Header Row",

        "Bill / Invoice Column",

        "Table Start",

        "Table End"
    ]

    rows = []

    for file_data in indexed_files:

        for table in file_data[
            "tables"
        ]:

            rows.append([

                table["filename"],

                table["sheet"],

                table["header"],

                table["header_row"],

                table["invoice_column"],

                table["start_column"],

                table["end_column"]
            ])

    if rows:

        display_excel_table(
            headers,
            rows,
            height=450
        )


# ============================================================
# DOWNLOAD EXCEL
# ============================================================

def create_download_excel(
    headers,
    rows
):

    output = io.BytesIO()

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = (
        "Search Result"
    )

    headers = list(
        headers
    )

    maximum_columns = len(
        headers
    )

    for row in rows:

        maximum_columns = max(
            maximum_columns,
            len(row)
        )

    while len(
        headers
    ) < maximum_columns:

        headers.append(
            f"Column {len(headers) + 1}"
        )

    # --------------------------------------------------------
    # Header
    # --------------------------------------------------------

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        worksheet.cell(
            row=1,
            column=column_number,
            value=(
                ""
                if is_blank(header)
                else header
            )
        )

    # --------------------------------------------------------
    # Data
    # --------------------------------------------------------

    for row_number, row in enumerate(
        rows,
        start=2
    ):

        row = list(
            row
        )

        if len(
            row
        ) < maximum_columns:

            row.extend(

                [None] * (
                    maximum_columns -
                    len(row)
                )
            )

        for column_number, value in enumerate(
            row[:maximum_columns],
            start=1
        ):

            worksheet.cell(
                row=row_number,
                column=column_number,
                value=(
                    None
                    if is_blank(value)
                    else value
                )
            )

    # --------------------------------------------------------
    # Freeze header
    # --------------------------------------------------------

    worksheet.freeze_panes = (
        "A2"
    )

    # --------------------------------------------------------
    # Auto width
    # --------------------------------------------------------

    for column_cells in worksheet.columns:

        maximum_length = 0

        for cell in column_cells:

            if cell.value is not None:

                maximum_length = max(
                    maximum_length,
                    len(
                        str(
                            cell.value
                        )
                    )
                )

        width = min(
            max(
                maximum_length + 2,
                10
            ),
            50
        )

        worksheet.column_dimensions[
            column_cells[
                0
            ].column_letter
        ].width = width

    workbook.save(
        output
    )

    output.seek(0)

    return output.getvalue()


# ============================================================
# DOWNLOAD FILE NAME
# ============================================================

def make_download_filename(
    filename,
    search_value
):

    base = re.sub(
        r"\.(xlsx|xlsm)$",
        "",
        filename,
        flags=re.IGNORECASE
    )

    base = re.sub(
        r'[\\/:*?"<>|]+',
        "_",
        base
    )

    search_part = re.sub(
        r'[\\/:*?"<>|]+',
        "-",
        search_value
    )

    return (
        f"{base}_"
        f"{search_part}_"
        f"Search_Result.xlsx"
    )


# ============================================================
# DISPLAY SEARCH RESULTS
# ============================================================

def display_results(
    results,
    search_value
):

    grouped = defaultdict(
        list
    )

    for result in results:

        key = (

            result["filename"],

            result["sheet"],

            result["header"],

            result["header_row"],

            result["invoice_column"],

            result["start_column"],

            result["end_column"]
        )

        grouped[
            key
        ].append(
            result
        )

    for group_number, (
        key,
        group
    ) in enumerate(
        grouped.items()
    ):

        (
            filename,
            sheet_name,
            invoice_header,
            header_row,
            invoice_column,
            start_column,
            end_column
        ) = key

        st.markdown(
            "---"
        )

        st.markdown(
            f"### 📄 {filename}"
        )

        st.write(
            f"**Sheet:** {sheet_name}"
        )

        st.write(
            f"**Bill / Invoice Header:** "
            f"`{invoice_header}`"
        )

        st.write(
            f"**Excel Header Row:** "
            f"{header_row}"
            f"  |  "
            f"**Bill / Invoice Column:** "
            f"{invoice_column}"
        )

        st.caption(
            f"Table columns: "
            f"{start_column} → {end_column}"
        )

        st.caption(
            f"Showing {len(group)} "
            f"complete matching Excel row(s)."
        )

        headers = list(
            group[0]["headers"]
        )

        rows = [

            result["row"]

            for result in group
        ]

        display_excel_table(
            headers,
            rows,
            height=650
        )

        download_data = create_download_excel(
            headers,
            rows
        )

        download_name = make_download_filename(
            filename,
            search_value
        )

        st.download_button(

            "⬇️ Download Table",

            data=download_data,

            file_name=download_name,

            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),

            key=(
                "download_"
                + str(
                    abs(
                        hash(
                            (
                                filename,
                                sheet_name,
                                invoice_header,
                                header_row,
                                invoice_column,
                                start_column,
                                end_column,
                                search_value,
                                group_number
                            )
                        )
                    )
                )
            )
        )


# ============================================================
# UPLOAD
# ============================================================

uploaded_files = st.file_uploader(

    "📂 Upload Excel files",

    type=[
        "xlsx",
        "xlsm"
    ],

    accept_multiple_files=True
)


# ============================================================
# APPLICATION
# ============================================================

if uploaded_files:

    st.success(
        f"✅ {len(uploaded_files)} "
        f"Excel file(s) uploaded"
    )

    # --------------------------------------------------------
    # Store files in memory
    # --------------------------------------------------------

    file_data = tuple(

        (
            file.name,
            file.getvalue()
        )

        for file in uploaded_files
    )

    # --------------------------------------------------------
    # Build index
    # --------------------------------------------------------

    with st.spinner(
        "📖 Reading Excel files..."
    ):

        indexed_files = build_search_index(
            file_data
        )

    # --------------------------------------------------------
    # Errors
    # --------------------------------------------------------

    for file_result in indexed_files:

        if file_result[
            "error"
        ]:

            st.error(
                f"❌ "
                f"{file_result['filename']}: "
                f"{file_result['error']}"
            )

    # --------------------------------------------------------
    # Count tables
    # --------------------------------------------------------

    total_tables = sum(

        len(
            item["tables"]
        )

        for item in indexed_files
    )

    if total_tables:

        st.success(
            f"✅ {total_tables} "
            f"Bill / Invoice table(s) detected"
        )

        with st.expander(
            "📋 View detected Bill / Invoice headers"
        ):

            display_detected_headers(
                indexed_files
            )

    else:

        st.warning(
            "⚠️ No Bill / Invoice headers detected."
        )

    # ========================================================
    # SEARCH AREA
    # ========================================================

    st.markdown(
        "## 🔎 Search Bill / Invoice Number"
    )

    search_value = st.text_input(

        "Bill / Invoice Number",

        placeholder=(
            "Example: 74, 162, "
            "CRSCL/19-20/1086, "
            "SMS/0504/2021-22"
        ),

        label_visibility="collapsed"
    )

    search_button = st.button(

        "🔎 SEARCH",

        type="primary",

        use_container_width=True
    )

    # ========================================================
    # SEARCH
    # ========================================================

    if search_button:

        if not search_value.strip():

            st.warning(
                "⚠️ Please enter a "
                "Bill / Invoice number."
            )

        elif total_tables == 0:

            st.error(
                "❌ No Bill / Invoice "
                "columns detected."
            )

        else:

            with st.spinner(
                "🔍 Searching..."
            ):

                results = search_excel(
                    indexed_files,
                    search_value
                )

            # ------------------------------------------------
            # FOUND
            # ------------------------------------------------

            if results:

                st.success(
                    f"✅ Found "
                    f"{len(results)} "
                    f"exact match(es)"
                )

                display_results(
                    results,
                    search_value
                )

            # ------------------------------------------------
            # NOT FOUND
            # ------------------------------------------------

            else:

                st.error(
                    f'❌ No exact match found for '
                    f'"{search_value}"'
                )

                st.caption(
                    "Search is performed only in "
                    "detected Bill / Invoice columns."
                )

else:

    st.info(
        "👆 Upload one or more Excel files to begin."
    )
